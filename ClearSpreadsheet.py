from openpyxl import Workbook, load_workbook

wb = load_workbook("DataSpread/data.xlsx")
ws = wb.active
initialRow = 2
currentRow = 2
openRow = False
while not openRow:
    if str(ws.cell(currentRow, 1).value) == "None":
        openRow = True
    else:
        currentRow += 1
print("Deleting a total of {} rows".format(currentRow - 2))
ws.delete_rows(initialRow, currentRow)
wb.save("DataSpread/data.xlsx")
print("Done!") 