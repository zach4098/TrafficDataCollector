import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def ReadTime(text):
    time = text.split('-')[1].split(' ')[3].split(':')
    return time
def TimeDifference(timeInit, timeFinal, vehicleCount, leftCount, rightCount):
    initHour = int(timeInit[0])
    initMinute = int(timeInit[1])
    initSecond = int(timeInit[2])

    finalHour = int(timeFinal[0])
    finalMinute = int(timeFinal[1])
    finalSecond = int(timeFinal[2])

    differenceInHours = finalHour-initHour
    differenceInMinutes = finalMinute - initMinute
    differenceInSeconds = finalSecond - initSecond

    if differenceInMinutes < 0:
        differenceInHours -= 1
        differenceInMinutes += 60
    if differenceInSeconds < 0:
        differenceInMinutes -= 1
        differenceInSeconds += 60
    
    totalHours = differenceInHours + (differenceInMinutes/60) + (differenceInSeconds/3600)
    totalHours = round(totalHours, 5)
    
    VperHour = round(vehicleCount/totalHours)
    VperHrL = round(leftCount/totalHours)
    VperHrR = round(rightCount/totalHours)


    return differenceInHours, differenceInMinutes, differenceInSeconds, totalHours, VperHour, VperHrL, VperHrR
def CountLeftRight(List):
    leftCount = 0
    rightCount = 0
    nullCount = 0
    for item in List:
        direction = item.split('-')[2]
        if direction == "left":
            leftCount += 1
        elif direction == "right":
            rightCount += 1
        else:
            nullCount += 1
    return leftCount, rightCount, nullCount
def ReadFile(folder, file):
    
    with open("{}/{}".format(folder, file), "r") as f:
            lines = f.readlines()
    count = 0
    for item in lines:
        item = item.replace("\n", "")
        item = item.replace("Vehicle ", "")
        lines[count] = item
        count += 1
    totalVehicles = count
    totalLeft, totalRight, totalNull = CountLeftRight(lines)
    lineInit = lines[0]
    lineFinal = lines[count - 1]
    timeInit = ReadTime(lineInit)
    timeFinal = ReadTime(lineFinal)
    differenceHour, differenceMinute, differenceSecond, totalHours, vPerHour, vPerHourL, vPerHourR = TimeDifference(timeInit, timeFinal, totalVehicles, totalLeft, totalRight)

    wb = load_workbook("DataSpread/data.xlsx")
    ws = wb.active
    currentRow = 2
    openRow = False
    while not openRow:
        if str(ws.cell(currentRow, 1).value) == "None":
            openRow = True
        else:
            currentRow += 1
    ws.cell(currentRow, 1).value = file
    ws.cell(currentRow, 2).value = "{}:{}:{}".format(timeInit[0], timeInit[1], timeInit[2])
    ws.cell(currentRow, 3).value = "{}:{}:{}".format(timeFinal[0], timeFinal[1], timeFinal[2])
    ws.cell(currentRow, 4).value = "{}:{}:{}".format(differenceHour, differenceMinute, differenceSecond)
    ws.cell(currentRow, 5).value = totalVehicles
    ws.cell(currentRow, 6).value = totalLeft
    ws.cell(currentRow, 7).value = totalRight
    ws.cell(currentRow, 8).value = totalNull
    ws.cell(currentRow, 9).value = totalHours
    ws.cell(currentRow, 10).value = vPerHour
    ws.cell(currentRow, 11).value = vPerHourL
    ws.cell(currentRow, 12).value = vPerHourR

    wb.save("DataSpread/data.xlsx")

folder = input("Select Folder: ")
collections = os.listdir("{}/".format(folder))

msg = "Select File to Read:\n"
count = 0
for item in collections:
    msg = msg + "{}: {}".format(count, str(item)) + "\n"
    count += 1
print(msg)

fileInput = input("Select File: ")

if fileInput == "all":
    count = 0
    for i in collections:
        ReadFile(folder, collections[count])
        count += 1
    print("Added a total of {} datasets!".format(count))
else:
    file = collections[int(fileInput)]
    ReadFile(folder, file)
    print("Done!")


