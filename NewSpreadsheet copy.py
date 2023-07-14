from openpyxl import Workbook, load_workbook

wb = Workbook()
workbookName = input("Enter Spreadsheet Name: ")
if workbookName == "":
    workbookName = "data"
wb.save("DataSpread/{}.xlsx".format(workbookName))
ws = wb.active
ws.cell(1, 1).value = "File Name"
ws.cell(1, 2).value = "Time Initial"
ws.cell(1, 3).value = "Time Final"
ws.cell(1, 4).value = "Time Total"
ws.cell(1, 5).value = "Total Vehicles"
ws.cell(1, 6).value = "Vehicles from Left"
ws.cell(1, 7).value = "Vehicles from Right"
ws.cell(1, 8).value = "Vehicles from N/A"
ws.cell(1, 9).value = "Total Hours"
ws.cell(1, 10).value = "Vehicles per Hour"
ws.cell(1, 11).value = "Left Vehicles per Hour"
ws.cell(1, 12).value = "Right Vehicles per Hour"

wb.save("DataSpread/data.xlsx")